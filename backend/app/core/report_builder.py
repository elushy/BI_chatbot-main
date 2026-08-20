import io
import json
import base64
import os
from typing import Dict, Any, List, Optional
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as OpenpyxlImage

from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Preformatted, Image as RLImage
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

# Determine font names supporting Turkish characters
FONT_REGULAR = 'Helvetica'
FONT_BOLD = 'Helvetica-Bold'
FONT_MONO = 'Courier'

try:
    # 1. Gather all potential font directories
    font_dirs = []
    
    # Windows paths
    windir = os.environ.get('WINDIR', 'C:\\Windows')
    font_dirs.append(os.path.join(windir, 'Fonts'))
    
    # Linux paths
    font_dirs.extend([
        '/usr/share/fonts',
        '/usr/share/fonts/truetype',
        '/usr/share/fonts/truetype/dejavu',
        '/usr/share/fonts/truetype/liberation',
        '/usr/share/fonts/truetype/ubuntu',
        '/usr/local/share/fonts',
        os.path.expanduser('~/.fonts')
    ])
    
    # 2. Candidate specifications: (name, regular_filename, bold_filename, mono_filename)
    candidates = [
        # Windows candidates
        ('Arial', 'arial.ttf', 'arialbd.ttf', 'cour.ttf'),
        ('Segoe UI', 'segoeui.ttf', 'segoeuib.ttf', 'cour.ttf'),
        # Linux DejaVu
        ('DejaVuSans', 'DejaVuSans.ttf', 'DejaVuSans-Bold.ttf', 'DejaVuSansMono.ttf'),
        # Linux Liberation
        ('LiberationSans', 'LiberationSans-Regular.ttf', 'LiberationSans-Bold.ttf', 'LiberationMono-Regular.ttf'),
        # Linux Ubuntu
        ('Ubuntu', 'Ubuntu-R.ttf', 'Ubuntu-B.ttf', 'UbuntuMono-R.ttf'),
    ]
    
    registered = False
    
    for name, reg_file, bold_file, mono_file in candidates:
        for f_dir in font_dirs:
            if not os.path.exists(f_dir):
                continue
            
            # Direct check
            reg_path = os.path.join(f_dir, reg_file)
            bold_path = os.path.join(f_dir, bold_file)
            mono_path = os.path.join(f_dir, mono_file)
            
            # If not found directly, check subdirectories recursively
            if not (os.path.exists(reg_path) and os.path.exists(bold_path)):
                found_reg, found_bold, found_mono = None, None, None
                for root, _, files in os.walk(f_dir):
                    for fn in files:
                        if fn.lower() == reg_file.lower():
                            found_reg = os.path.join(root, fn)
                        elif fn.lower() == bold_file.lower():
                            found_bold = os.path.join(root, fn)
                        elif fn.lower() == mono_file.lower():
                            found_mono = os.path.join(root, fn)
                    if found_reg and found_bold:
                        reg_path, bold_path = found_reg, found_bold
                        if found_mono:
                            mono_path = found_mono
                        break
            
            # Register if found
            if os.path.exists(reg_path) and os.path.exists(bold_path):
                pdfmetrics.registerFont(TTFont(name, reg_path))
                pdfmetrics.registerFont(TTFont(f"{name}-Bold", bold_path))
                FONT_REGULAR = name
                FONT_BOLD = f"{name}-Bold"
                
                if os.path.exists(mono_path):
                    pdfmetrics.registerFont(TTFont(f"{name}-Mono", mono_path))
                    FONT_MONO = f"{name}-Mono"
                else:
                    # Monospace fallback if custom TTF is not found
                    cour_path = os.path.join(os.path.join(windir, 'Fonts'), 'cour.ttf')
                    if os.path.exists(cour_path):
                        pdfmetrics.registerFont(TTFont('CourierNew', cour_path))
                        FONT_MONO = 'CourierNew'
                
                registered = True
                break
        if registered:
            break
            
except Exception as e:
    import sys
    print(f"Turkish font registration warning: {e}", file=sys.stderr)


def build_excel_report(
    columns: List[str],
    rows: List[List[Any]],
    title: str = "DeepBI Analiz Raporu",
    chart_image: Optional[str] = None
) -> io.BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Analiz Sonucu"
    
    # Enable grid lines explicitly
    ws.views.sheetView[0].showGridLines = True
    
    # Styling definitions
    header_fill = PatternFill(start_color="1F3D6B", end_color="1F3D6B", fill_type="solid")
    header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    
    zebra_fill = PatternFill(start_color="F6F8FA", end_color="F6F8FA", fill_type="solid")
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    
    border_side = Side(border_style="thin", color="D0D7DE")
    data_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
    
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")
    
    # Write Title
    ws.merge_cells("A1:E1")
    ws["A1"] = title
    ws["A1"].font = Font(name="Segoe UI", size=16, bold=True, color="1F3D6B")
    ws["A1"].alignment = align_left
    ws.row_dimensions[1].height = 35
    
    # Spacer row
    ws.row_dimensions[2].height = 10
    
    # Write Headers
    header_row = 3
    ws.row_dimensions[header_row].height = 28
    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=header_row, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = align_center
        cell.border = data_border
        
    # Write Data
    start_row = 4
    
    is_anomaly_report = "Durum" in columns
    is_correlation_report = len(columns) > 1 and columns[0] == "Değişken"
    
    durum_col_idx = columns.index("Durum") + 1 if is_anomaly_report else None
    
    anomaly_fill = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")
    pos_fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
    neg_fill = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")
    mod_fill = PatternFill(start_color="FFF8E1", end_color="FFF8E1", fill_type="solid")
    
    for r_idx, row_data in enumerate(rows):
        current_row = start_row + r_idx
        ws.row_dimensions[current_row].height = 20
        
        row_is_anomaly = False
        if is_anomaly_report and durum_col_idx is not None:
            try:
                row_is_anomaly = (row_data[durum_col_idx - 1] == 'Anomali')
            except Exception:
                pass
                
        base_fill = zebra_fill if r_idx % 2 == 1 else white_fill
        if row_is_anomaly:
            base_fill = anomaly_fill
            
        for c_idx, cell_value in enumerate(row_data, 1):
            val_to_write = cell_value
            if cell_value is not None and not isinstance(cell_value, (int, float, bool, str)):
                val_to_write = str(cell_value)
                
            cell = ws.cell(row=current_row, column=c_idx, value=val_to_write)
            cell.border = data_border
            cell.font = Font(name="Segoe UI", size=10)
            
            # Apply cell-specific conditional formatting
            cell_fill = base_fill
            if is_correlation_report and c_idx > 1:
                if isinstance(cell_value, (int, float)):
                    val = float(cell_value)
                    if val >= 0.5:
                        cell_fill = pos_fill
                    elif val <= -0.5:
                        cell_fill = neg_fill
                    elif 0.3 <= abs(val) < 0.5:
                        cell_fill = mod_fill
                        
            cell.fill = cell_fill
            
            # Align based on data type
            if isinstance(cell_value, (int, float)):
                cell.alignment = align_right
                if isinstance(cell_value, float):
                    cell.number_format = "0.00" if is_correlation_report else "#,##0.00"
                else:
                    cell.number_format = "#,##0"
            elif isinstance(cell_value, bool):
                cell.alignment = align_center
            else:
                cell.alignment = align_left
                
    # Auto-adjust column widths
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.row == 1:
                continue # ignore title row length
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
        
    # Write Chart if provided
    if chart_image:
        try:
            if "," in chart_image:
                header, encoded = chart_image.split(",", 1)
            else:
                encoded = chart_image
            image_data = base64.b64decode(encoded)
            img_buf = io.BytesIO(image_data)
            
            img = OpenpyxlImage(img_buf)
            img.width = 600
            img.height = 300
            
            # Side-by-side if column count is low, stacked below if column count is high
            if len(columns) <= 5:
                ws.add_image(img, 'G3')
            else:
                chart_row = start_row + len(rows) + 3
                ws.add_image(img, f'A{chart_row}')
        except Exception as e:
            print(f"Failed to add chart to Excel report: {e}")
            
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def escape_for_paragraph(text: str) -> str:
    if not text:
        return ""
    text = text.replace("&", "&amp;")
    text = text.replace("<", "&lt;")
    text = text.replace(">", "&gt;")
    text = text.replace("&lt;b&gt;", "<b>").replace("&lt;/b&gt;", "</b>")
    text = text.replace("&lt;i&gt;", "<i>").replace("&lt;/i&gt;", "</i>")
    text = text.replace("&lt;u&gt;", "<u>").replace("&lt;/u&gt;", "</u>")
    return text


def build_pdf_report(
    question: str,
    summary_text: str,
    code: str,
    code_language: str,
    columns: List[str],
    rows: List[List[Any]],
    session_title: str = "DeepBI Sohbet Oturumu",
    chart_image: Optional[str] = None
) -> io.BytesIO:
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=letter,
        rightMargin=40,
        leftMargin=40,
        topMargin=40,
        bottomMargin=40
    )
    
    styles = getSampleStyleSheet()
    
    # Premium Typography with Turkish Font support
    title_style = ParagraphStyle(
        'DocTitle',
        parent=styles['Normal'],
        fontName=FONT_BOLD,
        fontSize=20,
        textColor=colors.HexColor('#1f3d6b'),
        spaceAfter=6
    )
    
    subtitle_style = ParagraphStyle(
        'DocSubtitle',
        parent=styles['Normal'],
        fontName=FONT_REGULAR,
        fontSize=10,
        textColor=colors.HexColor('#57606a'),
        spaceAfter=20
    )
    
    h2_style = ParagraphStyle(
        'SectionHeading',
        parent=styles['Heading2'],
        fontName=FONT_BOLD,
        fontSize=12,
        textColor=colors.HexColor('#24292f'),
        spaceBefore=14,
        spaceAfter=6,
        keepWithNext=True
    )
    
    body_style = ParagraphStyle(
        'Body',
        parent=styles['BodyText'],
        fontName=FONT_REGULAR,
        fontSize=10,
        textColor=colors.HexColor('#24292f'),
        leading=14,
        spaceAfter=10
    )
    
    code_style = ParagraphStyle(
        'CodeText',
        parent=styles['Normal'],
        fontName=FONT_MONO,
        fontSize=9,
        textColor=colors.HexColor('#e6edf3'),
        leading=12
    )
    
    # Table Column/Cell Wrap ParagraphStyle
    th_style = ParagraphStyle(
        'TableHeader',
        parent=styles['Normal'],
        fontName=FONT_BOLD,
        fontSize=9,
        textColor=colors.white,
        alignment=TA_CENTER
    )
    
    td_left_style = ParagraphStyle(
        'TableDataLeft',
        parent=styles['Normal'],
        fontName=FONT_REGULAR,
        fontSize=9,
        textColor=colors.HexColor('#24292f'),
        alignment=TA_LEFT
    )

    td_right_style = ParagraphStyle(
        'TableDataRight',
        parent=styles['Normal'],
        fontName=FONT_REGULAR,
        fontSize=9,
        textColor=colors.HexColor('#24292f'),
        alignment=TA_RIGHT
    )
    
    story = []
    
    # 1. Header Block
    story.append(Paragraph("DeepBI Analiz ve Yönetici Raporu", title_style))
    story.append(Paragraph(f"Oturum: {session_title} | Raporlama Tarihi: Güncel", subtitle_style))
    story.append(Spacer(1, 10))
    
    # 2. User Question Block
    story.append(Paragraph("Sorgulanan Metrik / Soru", h2_style))
    question_box_content = f"<b>Soru:</b> {escape_for_paragraph(question)}"
    story.append(Paragraph(question_box_content, body_style))
    story.append(Spacer(1, 10))
    
    # 3. Code Block
    if code:
        story.append(Paragraph(f"Üretilen Analiz Kodu ({code_language.upper()})", h2_style))
        # Monospace formatted block
        code_p = Preformatted(code, code_style)
        
        # Wrap code in a table for background color styling
        code_container = Table([[code_p]], colWidths=[530])
        code_container.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,-1), colors.HexColor('#161b22')),
            ('ALIGN', (0,0), (-1,-1), 'LEFT'),
            ('VALIGN', (0,0), (-1,-1), 'TOP'),
            ('BOTTOMPADDING', (0,0), (-1,-1), 10),
            ('TOPPADDING', (0,0), (-1,-1), 10),
            ('LEFTPADDING', (0,0), (-1,-1), 10),
            ('RIGHTPADDING', (0,0), (-1,-1), 10),
            ('BOX', (0,0), (-1,-1), 1, colors.HexColor('#30363d')),
        ]))
        story.append(code_container)
        story.append(Spacer(1, 12))
        
    # 4. Summary / Insights Section
    if summary_text:
        story.append(Paragraph("Yapay Zekâ Bulguları & Yorumlar", h2_style))
        clean_text = summary_text.replace("### ", "").replace("## ", "").replace("**", "")
        clean_text = escape_for_paragraph(clean_text)
        story.append(Paragraph(clean_text, body_style))
        story.append(Spacer(1, 10))
        
    # 4.5. Chart Section if provided
    if chart_image:
        try:
            if "," in chart_image:
                header, encoded = chart_image.split(",", 1)
            else:
                encoded = chart_image
                
            image_data = base64.b64decode(encoded)
            img_buf = io.BytesIO(image_data)
            
            # 500x250 fits beautifully on standard Letter size page
            rl_img = RLImage(img_buf, width=500, height=250)
            story.append(Paragraph("Grafik Görselleştirmesi", h2_style))
            story.append(rl_img)
            story.append(Spacer(1, 12))
        except Exception as e:
            print(f"Failed to add chart to PDF report: {e}")
            
    # 5. Result Table
    if columns and rows:
        story.append(Paragraph("Detaylı Bulgular Tablosu", h2_style))
        
        # Max limit table row/col size in PDF to prevent page break overflows
        max_rows = min(len(rows), 40)
        max_cols = min(len(columns), 8)
        
        subset_cols = columns[:max_cols]
        subset_rows = [r[:max_cols] for r in rows[:max_rows]]
        
        # Build Table Data Flowables
        table_data = []
        
        # Header Flowable Row
        header_flow_row = [Paragraph(escape_for_paragraph(c), th_style) for c in subset_cols]
        table_data.append(header_flow_row)
        
        # Data Flowable Rows
        for r_idx, row in enumerate(subset_rows):
            flow_row = []
            for cell_val in row:
                cell_str = "null" if cell_val is None else str(cell_val)
                is_num = isinstance(cell_val, (int, float))
                style_to_use = td_right_style if is_num else td_left_style
                flow_row.append(Paragraph(escape_for_paragraph(cell_str), style_to_use))
            table_data.append(flow_row)
            
        # Dynamically distribute width
        col_width = 530 / max_cols
        col_widths = [col_width] * max_cols
        
        pdf_table = Table(table_data, colWidths=col_widths)
        
        # Build styling logic
        t_style = [
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1f3d6b')),
            ('ALIGN', (0,0), (-1,-1), 'LEFT'),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#d0d7de')),
            ('TOPPADDING', (0,0), (-1,-1), 5),
            ('BOTTOMPADDING', (0,0), (-1,-1), 5),
        ]
        
        # Alternating zebra fills
        for idx in range(1, len(table_data)):
            bg_color = colors.HexColor('#f6f8fa') if idx % 2 == 1 else colors.white
            t_style.append(('BACKGROUND', (0, idx), (-1, idx), bg_color))
            
        pdf_table.setStyle(TableStyle(t_style))
        story.append(pdf_table)
        
        if len(rows) > max_rows:
            story.append(Spacer(1, 5))
            story.append(Paragraph(f"<i>* Not: Bu rapor analiz edilen {len(rows)} satır veriden ilk {max_rows} satırını göstermektedir.</i>", subtitle_style))
            
    doc.build(story)
    buffer.seek(0)
    return buffer
